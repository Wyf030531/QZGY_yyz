# -*- coding: utf-8 -*-
"""
根据 Excel 植被类型代码表，对 veg100w.tif 进行重分类
重分类规则：
1 = 针叶类（针叶林 + 针叶灌丛）
2 = 针阔叶混交林
3 = 阔叶类（阔叶林 + 阔叶灌丛）
4 = 草地（草丛、草原、草甸合并）
6 = 高山植被
7 = 栽培植被
0 = 其他（荒漠、沼泽、沙漠、其它未匹配类型）
255 = NoData

更新规则：凡属于“2针阔叶混交林”的像元重分类为 2，属于“1针叶林”的重分类为 1，属于“3阔叶林”的重分类为 3，
属于“7草丛”“6草原”或“8草甸”的统一合并重分类为 4；属于“4灌丛”时，不再统一重分类为单独灌丛类，而是根据“植被亚类”进一步并入针叶或阔叶两类，
其中亚类为“(24)亚高山常绿针叶灌丛”的重分类为 1，亚类为“(18)温带落叶灌丛”“(19)亚热带、热带常绿阔叶、落叶阔叶灌丛(常含稀树)”“(21)亚热带、热带旱生常绿肉质多刺灌丛”“(22)亚高山落叶阔叶灌丛”“(23)亚高山硬叶常绿阔叶灌丛”的重分类为 3；
属于“10高山植被”的重分类为 6，属于“11栽培植被”的重分类为 7，而属于“5荒漠”“9沼泽”以及名称中含“其他”的类型统一重分类为 0，未匹配到的类型也默认归为 0。
"""

import os
import numpy as np
import rasterio
import openpyxl


# =========================
# 1. 输入输出路径
# =========================
tif_path = r"F:\wyf\数据存储\组内数据\严艳梓老师合作论文相关\veg\222\veg100w.tif"
excel_path = r"F:\wyf\数据存储\组内数据\严艳梓老师合作论文相关\veg\植被类型代码表.xlsx"
out_tif = r"F:\wyf\数据存储\组内数据\严艳梓老师合作论文相关\veg\222\result\veg100w_reclass.tif"


# =========================
# 2. 建立重分类映射
# =========================
def normalize_text(x):
    """统一清洗文本，避免空格和 None 影响判断"""
    if x is None:
        return ""
    return str(x).strip().replace(" ", "").replace("（", "(").replace("）", ")")


def build_reclass_map_from_excel(excel_file):
    """
    从 Excel 中读取 Value、植被大类、植被亚类 的对应关系，建立重分类字典
    返回：
        reclass_map: {原始value: 新类别值}
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]

    # 查找字段列
    try:
        value_idx = headers.index("Value")
        big_class_idx = headers.index("植被大类")
        sub_class_idx = headers.index("植被亚类")
    except ValueError:
        raise ValueError("Excel 中未找到 'Value'、'植被大类' 或 '植被亚类' 列，请检查表头名称。")

    reclass_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_value = row[value_idx]
        big_class = normalize_text(row[big_class_idx])
        sub_class = normalize_text(row[sub_class_idx])

        if raw_value is None:
            continue

        try:
            raw_value = int(raw_value)
        except Exception:
            continue

        # =========================
        # 按更新后的规则重分类
        # =========================

        # 1. 针阔叶混交林
        if "2针阔叶混交林" in big_class:
            new_class = 2

        # 2. 针叶林
        elif "1针叶林" in big_class:
            new_class = 1

        # 3. 阔叶林
        elif "3阔叶林" in big_class:
            new_class = 3

        # 4. 草地：草丛、草原、草甸
        elif ("7草丛" in big_class) or ("6草原" in big_class) or ("8草甸" in big_class):
            new_class = 4

        # 5. 灌丛：不再单独归为 5，而是看亚类并入针叶/阔叶
        elif "4灌丛" in big_class:
            # 针叶灌丛 -> 并入针叶类
            if ("(24)" in sub_class) or ("24亚高山常绿针叶灌丛" in sub_class):
                new_class = 1

            # 阔叶灌丛 -> 并入阔叶类
            elif (
                ("(18)" in sub_class) or ("18温带落叶灌丛" in sub_class) or
                ("(19)" in sub_class) or ("19亚热带、热带常绿阔叶、落叶阔叶灌丛(常含稀树)" in sub_class) or
                ("(21)" in sub_class) or ("21亚热带、热带旱生常绿肉质多刺灌丛" in sub_class) or
                ("(22)" in sub_class) or ("22亚高山落叶阔叶灌丛" in sub_class) or
                ("(23)" in sub_class) or ("23亚高山硬叶常绿阔叶灌丛" in sub_class)
            ):
                new_class = 3

            # 若是灌丛但亚类没识别到，默认仍归入灌丛主体更接近的阔叶类
            # 也可改成 0，看你需求
            else:
                new_class = 3

        # 6. 高山植被
        elif "10高山植被" in big_class:
            new_class = 6

        # 7. 栽培植被
        elif "11栽培植被" in big_class:
            new_class = 7

        # 0. 其他：荒漠、沼泽、沙漠、其它未匹配
        elif ("5荒漠" in big_class) or ("9沼泽" in big_class) or ("其他" in big_class):
            new_class = 0

        else:
            # Excel 中若有未列出的类型，默认归为其他
            new_class = 0

        reclass_map[raw_value] = new_class

    return reclass_map


# =========================
# 3. 栅格重分类
# =========================
def reclassify_tif(in_tif, out_tif, reclass_map, out_nodata=255):
    with rasterio.open(in_tif) as src:
        data = src.read(1)
        profile = src.profile.copy()
        src_nodata = src.nodata

        # 默认全部赋值为 0（其他）
        out_arr = np.zeros(data.shape, dtype=np.uint8)

        # 识别输入 NoData
        if src_nodata is not None:
            nodata_mask = (data == src_nodata)
        else:
            nodata_mask = np.zeros(data.shape, dtype=bool)

        valid_mask = ~nodata_mask
        unique_vals = np.unique(data[valid_mask])

        # 逐个原始值重分类
        for val in unique_vals:
            new_val = reclass_map.get(int(val), 0)
            out_arr[data == val] = new_val

        # 恢复 NoData
        out_arr[nodata_mask] = out_nodata

        # 更新输出参数
        profile.update(
            dtype=rasterio.uint8,
            count=1,
            compress="lzw",
            nodata=out_nodata
        )

        # 创建输出目录
        out_dir = os.path.dirname(out_tif)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir)

        with rasterio.open(out_tif, "w", **profile) as dst:
            dst.write(out_arr, 1)


# =========================
# 4. 统计输出
# =========================
def print_statistics(out_tif, out_nodata=255):
    class_names = {
        0: "其他",
        1: "针叶类（含针叶林、针叶灌丛）",
        2: "针阔叶混交林",
        3: "阔叶类（含阔叶林、阔叶灌丛）",
        4: "草地",
        6: "高山植被",
        7: "栽培植被",
        255: "NoData"
    }

    with rasterio.open(out_tif) as ds:
        arr = ds.read(1)

    unique, counts = np.unique(arr, return_counts=True)

    print("\n重分类完成，像元统计如下：")
    for val, cnt in zip(unique, counts):
        name = class_names.get(int(val), "未知")
        print(f"{val} -> {name}: {cnt} 个像元")


# =========================
# 5. 主程序
# =========================
if __name__ == "__main__":
    print("开始读取 Excel 并建立重分类规则...")
    reclass_map = build_reclass_map_from_excel(excel_path)
    print(f"共读取 {len(reclass_map)} 条 Value 映射规则。")

    print("开始进行栅格重分类...")
    reclassify_tif(tif_path, out_tif, reclass_map, out_nodata=255)

    print(f"\n输出文件已生成：\n{out_tif}")
    print_statistics(out_tif, out_nodata=255)

    print("\n处理完成。")